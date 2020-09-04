#-*-coding:utf-8-*-
from openpyxl import load_workbook
import random
from tkinter import *
from tkinter import filedialog
import tkinter.font as tkFont


WORDNUM=0

class WordTest():
    
    word_list=[]
    
    def __init__(self):
        self.excelSetting()

        #window create
        self.window=Tk()
        
        self.windowSetting()
        self.widgetSetting()
        self.widgetPlace()
        
        #enter key event
        self.window.bind('<Return>',self.enterEevent) 
        #window run
        self.window.mainloop()

    def excelSetting(self):
        filename=filedialog.askopenfilename(initialdir="./",titl="공부할 단어장 선택",filetypes=(("xlsx files","*.xlsx"),("all files","*")))
        self.loadEx = load_workbook(filename, data_only=True)
        self.wordExl = self.loadEx['Sheet1']
        self.word_list=list(range(1,self.wordExl.max_row))
        random.shuffle(self.word_list)

        
    def windowSetting(self):
        self.window.title("단어 연습장")
        self.window.geometry("340x200")
        self.window.resizable(False,False)
        
    def widgetSetting(self):
        self.fontstyle=tkFont.Font(size=20)
        self.keybstyle=tkFont.Font(size=13)
        self.anwstyle=tkFont.Font(size=13)
        self.word=Label(self.window,text=self.wordExl.cell(self.word_list[WORDNUM],1).value,font=self.fontstyle)
        self.answer=Label(self.window,text=self.wordExl.cell(self.word_list[WORDNUM],2).value,font=self.anwstyle)
        self.anwB=Button(self.window,text="정답 확인", command=self.answerButton)
        self.nextB=Button(self.window, text="다음 단어", command=self.nextButton)
        self.inpKeb=Entry(self.window,width=15,font=self.keybstyle)

    def widgetPlace(self):
        self.word.pack(side="top")
        self.inpKeb.pack()
        self.answer.pack_forget()
        self.nextB.pack(side="bottom",padx=10)
        self.anwB.pack(side="bottom",padx=10)
        
    
    def answerButton(self):
        if self.anwB.cget("text") == "정답 확인":
            self.answer.pack()
            self.anwB.configure(text="정답 숨기기")
        else:
            self.answer.pack_forget()
            self.anwB.configure(text="정답 확인")
        
    def nextButton(self):
        global WORDNUM
        if WORDNUM==(self.wordExl.max_row-2):
            WORDNUM=-1
            random.shuffle(self.word_list)
            print("처음으로 돌아왔습니다, 단어를 다시 섞습니다.")
            
        self.answer.pack_forget()
        self.inpKeb.delete(0,END)
        
        self.anwB.configure(text="정답 확인")
        self.anwB.pack(side="bottom",padx=10)
        
        WORDNUM=WORDNUM+1
        self.word.configure(text=self.wordExl.cell(self.word_list[WORDNUM],1).value)
        self.answer.configure(text=self.wordExl.cell(self.word_list[WORDNUM],2).value)

    def enterEevent(self,event):
        self.answer.pack()
        self.anwB.pack_forget()
    
add=WordTest()


